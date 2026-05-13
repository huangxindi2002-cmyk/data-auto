"""
Export pipeline results to Excel.
Format mirrors data-updates/index.html downloadExcel():

Sheet "YYYY_MM":
  Row 0 : [blank, cat1, cat2, ...]          – category headers (bold)
  Row 1 : [blank, total1, total2, ...]      – totals (2 decimal places)
  Rows 2+: sparse – [appName, ..., time_in_col, ...]  (one row per app, null-padded)
           blank rows separate categories
  Last   : ["其他", others1, others2, ...]  – remainder row

Sheet "未分类App" (only if unknowns exist):
  Col A = app name, Col B = time (billion min)

Red bold font for: WhatsApp Messenger & WhatsApp Business, TikTok, Kwai
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import config

_RED_APPS = {"WhatsApp Messenger & WhatsApp Business", "TikTok", "Kwai"}
_RED_FONT = Font(color="FF0000", bold=True)
_BOLD_FONT = Font(bold=True)


def _col_letter(idx):
    """0-based column index → letter (A, B, ...)."""
    return openpyxl.utils.get_column_letter(idx + 1)


def export(pipeline_result, month, output_path):
    """
    pipeline_result: output of pipeline.run()
    month:           "2025-07"
    output_path:     path to write .xlsx
    """
    cats = config.CATEGORIES_ZH
    cat_col = {c: i + 1 for i, c in enumerate(cats)}  # 0-based col index (col A=0)
    result = pipeline_result["categories"]
    unknown = pipeline_result["unknown"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month.replace("-", "_")

    # ── Build app rows list (each entry: (app_name, cat, display_time) or None for blank) ──
    app_rows = []
    for cat in cats:
        for item in result[cat]["top5"]:
            app_rows.append((item["name"], cat, item["display_time"]))
        app_rows.append(None)  # blank separator

    total_data_rows = 2 + len(app_rows) + 1  # header + totals + app rows + others

    # ── Row 0: category headers ──
    for i, cat in enumerate(cats):
        cell = ws.cell(row=1, column=i + 2, value=cat)
        cell.font = _BOLD_FONT

    # ── Row 1: totals ──
    for i, cat in enumerate(cats):
        ws.cell(row=2, column=i + 2, value=round(result[cat]["total"], 2))

    # ── Rows 2+: app data ──
    for idx, entry in enumerate(app_rows):
        row_num = idx + 3  # 1-based, row 1=headers, row 2=totals, row 3=first app
        if entry is None:
            continue
        app_name, cat, display_time = entry
        name_cell = ws.cell(row=row_num, column=1, value=app_name)
        if app_name in _RED_APPS:
            name_cell.font = _RED_FONT
        col_idx = cat_col[cat]
        ws.cell(row=row_num, column=col_idx + 1, value=round(display_time, 3))

    # ── Last row: 其他 ──
    others_row = 3 + len(app_rows)
    ws.cell(row=others_row, column=1, value="其他")
    for i, cat in enumerate(cats):
        v = result[cat]["others"]
        if v > 0.001:
            ws.cell(row=others_row, column=i + 2, value=round(v, 3))

    # ── Column widths ──
    ws.column_dimensions["A"].width = 32
    for i in range(len(cats)):
        ws.column_dimensions[_col_letter(i + 1)].width = 14

    # ── Sheet 2: unknown apps ──
    if unknown:
        ws2 = wb.create_sheet("未分类App")
        ws2.append(["未分类 App", "时长（十亿分钟）"])
        ws2["A1"].font = _BOLD_FONT
        ws2["B1"].font = _BOLD_FONT
        for item in unknown:
            ws2.append([item["name"], round(item["time"], 3)])
        ws2.column_dimensions["A"].width = 50
        ws2.column_dimensions["B"].width = 20

    wb.save(output_path)
    print(f"Saved: {output_path}")
